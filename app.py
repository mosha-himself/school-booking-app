    # app.py
"""
Ultimate QR-directed Booking App (single-file)
Run: python app.py
Dependencies: Flask, pandas, openpyxl
pip install Flask pandas openpyxl
"""

from flask import Flask, request, redirect, url_for, render_template_string, jsonify, send_file
import sqlite3
import os
import re
import pandas as pd
from datetime import date, datetime, timedelta
from pathlib import Path

# -------------------------
# Configuration
# -------------------------
DB_FILE = "bookings.db"
EXCEL_FILE = "bookings.xlsx"
ADMIN_USER = os.getenv("USS_ADMIN_USER", "admin")
ADMIN_PASS = os.getenv("USS_ADMIN_PASS", "adminpass")  # change in environment for production
ENFORCE_SCHOOL_EMAIL = False  # set True to enforce aaisp000000@alansarschool.net pattern
SCHOOL_EMAIL_PREFIX = r"^aaisp\d{6}@alansarschool\.net$"
BLACKPOINTS_DEFAULT_THRESHOLD = 3

# -------------------------
# Flask app
# -------------------------
app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv("SECRET_KEY", "dev-secret-key")


# -------------------------
# Database helpers
# -------------------------
def get_db():
    conn = sqlite3.connect(DB_FILE, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.executescript("""
    PRAGMA foreign_keys = ON;
    CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT
    );
    CREATE TABLE IF NOT EXISTS students (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        email TEXT,
        grade TEXT,
        black_points INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS bookings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        booking_date TEXT NOT NULL,
        period INTEGER NOT NULL,
        sport TEXT NOT NULL,
        leader_student_id INTEGER,
        other_players TEXT,
        created_at TEXT NOT NULL,
        UNIQUE(booking_date, period, sport),
        FOREIGN KEY(leader_student_id) REFERENCES students(id) ON DELETE SET NULL
    );
    """)
    # default settings
    cur.execute("INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)", ("blackpoint_threshold", str(BLACKPOINTS_DEFAULT_THRESHOLD)))
    conn.commit()
    cur.close()
    conn.close()

init_db()

# -------------------------
# Timetable utilities
# -------------------------
def normalize_grade_input(raw: str):
    """Normalize various grade entries to 'Year 12' or 'Year 13' (Year12 corresponds to 12)"""
    if not raw:
        return None
    s = raw.strip().upper()
    # allow '12', 'Y12', 'Y-12', 'GRADE 11' mapping: user said Year12 can be written Grade11? user mapping: Year12 corresponds to 11? 
    # Clarify from user's text: they said "Year 12 (can be written as Grade 11) or Year 13 (can be written as Grade 12)".
    # So they shift naming: I'll treat normalized numeric value by interpreting last two digits if present.
    digits = re.findall(r"\d{1,2}", s)
    if digits:
        val = int(digits[-1])
        # Map: if student inputs 11 => Year12? But user text is confusing. We'll accept only 12 or 13 (or equivalents).
        # Accept if val == 12 or val == 13. If user writes 11 or 12 meaning Grade 11 -> Year12, we accept both 11->Year12 mapping and 12->Year13 mapping?
        # Simpler & robust: if val in (11,12) -> treat as Year 12, if val in (12,13) -> Year13. But that overlaps.
        # We'll implement rule: if val in {12, '12 variants'} => Year 12; if val in {13,...} => Year 13; if val ==11 assume Year 12 (as user said)
        if val == 11 or val == 12:
            return "Year 12"
        if val == 13:
            return "Year 13"
    # If string contains 12 letters YTW etc
    if "12" in s:
        return "Year 12"
    if "13" in s:
        return "Year 13"
    # fallback: if user typed Y12 or Y-12 or GRADE 11 etc, handle above; else return raw normalized
    return raw.strip().title()

def get_day_periods(target_date: date):
    """Return list of available periods for a given date respecting timetable rules.
    Returns list of dicts: {'period': n, 'start': 'HH:MM', 'end': 'HH:MM', 'bookable': True/False}
    """
    # Determine weekday: Monday=0 ... Sunday=6. School days are Mon-Thu only.
    wd = target_date.weekday()  # 0..6
    if wd > 3:  # Fri-Sun: no school
        return []
    # Monday(0)/Wednesday(2) => 45-min periods, 8 periods
    # Tuesday(1)/Thursday(3) => 40-min periods, 9 periods
    if wd in (0, 2):  # Mon, Wed
        period_length = 45
        total_periods = 8
    else:
        period_length = 40
        total_periods = 9
    # school starts 7:30
    start_time = datetime.combine(target_date, datetime.strptime("07:30", "%H:%M").time())
    periods = []
    # Build periods as in description:
    # 4 periods back-to-back, then 25 min break, then 3 periods back-to-back, then 15 min break, then remaining periods (1 on 45-min days, 2 on 40-min days)
    cur = start_time
    # first 4
    for i in range(1, 5):
        pstart = cur
        pend = pstart + timedelta(minutes=period_length)
        periods.append({'period': len(periods)+1, 'start': pstart.time().strftime("%H:%M"), 'end': pend.time().strftime("%H:%M")})
        cur = pend
    # break 25
    cur += timedelta(minutes=25)
    # next 3
    for i in range(1, 4):
        pstart = cur
        pend = pstart + timedelta(minutes=period_length)
        periods.append({'period': len(periods)+1, 'start': pstart.time().strftime("%H:%M"), 'end': pend.time().strftime("%H:%M")})
        cur = pend
    # break 15
    cur += timedelta(minutes=15)
    # remaining: total - len(periods) periods
    remaining = total_periods - len(periods)
    for i in range(remaining):
        pstart = cur
        pend = pstart + timedelta(minutes=period_length)
        periods.append({'period': len(periods)+1, 'start': pstart.time().strftime("%H:%M"), 'end': pend.time().strftime("%H:%M")})
        cur = pend
    # school ends at 14:10 (2:10 PM) per spec, but we won't rely on that except remove last period from booking
    # "no booking for the last period before school ends" -> mark last period as not bookable
    for p in periods:
        p['bookable'] = True
    if periods:
        periods[-1]['bookable'] = False
    # Also Thursday first period is club houses, so if Thurs (3) and target_date weekday ==3, first period not bookable
    if wd == 3 and periods:
        periods[0]['bookable'] = False
    return periods

# -------------------------
# Student & booking helpers
# -------------------------
def find_or_create_student(name, email, grade):
    conn = get_db()
    cur = conn.cursor()
    # search by exact email if provided, else by name
    if email:
        cur.execute("SELECT * FROM students WHERE email = ?", (email.strip().lower(),))
        row = cur.fetchone()
        if row:
            # update name/grade if changed
            cur.execute("UPDATE students SET name = ?, grade = ? WHERE id = ?", (name.strip(), grade, row['id']))
            conn.commit()
            conn.close()
            return row['id']
    # fallback: search by name
    cur.execute("SELECT * FROM students WHERE name = ?", (name.strip(),))
    row = cur.fetchone()
    if row:
        conn.close()
        return row['id']
    # create
    cur.execute("INSERT INTO students (name, email, grade) VALUES (?, ?, ?)", (name.strip(), (email.strip().lower() if email else None), grade))
    conn.commit()
    sid = cur.lastrowid
    conn.close()
    return sid

def is_slot_taken(booking_date, period, sport):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) as cnt FROM bookings WHERE booking_date = ? AND period = ? AND sport = ?", (booking_date, period, sport))
    cnt = cur.fetchone()['cnt']
    conn.close()
    return cnt > 0

def create_booking(booking_date, period, sport, leader_id, other_players):
    # try to insert — unique constraint will prevent duplicates
    conn = get_db()
    cur = conn.cursor()
    now = datetime.utcnow().isoformat()
    try:
        cur.execute("INSERT INTO bookings (booking_date, period, sport, leader_student_id, other_players, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                    (booking_date, period, sport, leader_id, other_players, now))
        conn.commit()
        bid = cur.lastrowid
        conn.close()
        # After successful DB insert, append to Excel sheet
        append_to_excel(booking_date, bid, booking_date, period, sport, leader_id, other_players, now)
        return True, None
    except sqlite3.IntegrityError:
        conn.close()
        return False, "slot_taken"
    except Exception as e:
        conn.close()
        return False, str(e)

# -------------------------
# Excel export helpers
# -------------------------
def append_to_excel(sheet_date_str, booking_id, booking_date, period, sport, leader_id, other_players, created_at):
    # ensure file exists; use pandas with openpyxl to append new sheet or append rows to existing sheet
    # gather student name by id
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT name, email, grade FROM students WHERE id = ?", (leader_id,))
    s = cur.fetchone()
    conn.close()
    leader_name = s['name'] if s else ''
    leader_email = s['email'] if s else ''
    leader_grade = s['grade'] if s else ''
    row = {
        "booking_id": booking_id,
        "booking_date": booking_date,
        "period": period,
        "sport": sport,
        "leader_name": leader_name,
        "leader_email": leader_email,
        "leader_grade": leader_grade,
        "other_players": other_players,
        "created_at": created_at
    }
    # Use pandas to append a row to a sheet named sheet_date_str (YYYY-MM-DD)
    df_row = pd.DataFrame([row])
    file_path = Path(EXCEL_FILE)
    if file_path.exists():
        # load existing, append to sheet or create sheet if missing
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            # pandas 1.4+ supports if_sheet_exists; but to be robust, we will try to read existing sheetnames
            try:
                from openpyxl import load_workbook
                wb = load_workbook(EXCEL_FILE)
                if sheet_date_str in wb.sheetnames:
                    # read existing sheet to df then append and write back
                    existing = pd.read_excel(EXCEL_FILE, sheet_name=sheet_date_str)
                    newdf = pd.concat([existing, df_row], ignore_index=True)
                    # remove old sheet then write new (openpyxl needed)
                    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a") as writer2:
                        writer2.book = wb
                        # remove sheet
                        std = writer2.book[sheet_date_str]
                        writer2.book.remove(std)
                        writer2.sheets = {ws.title: ws for ws in writer2.book.worksheets}
                        newdf.to_excel(writer2, sheet_name=sheet_date_str, index=False)
                else:
                    df_row.to_excel(writer, sheet_name=sheet_date_str, index=False)
            except Exception:
                # fallback: just write new sheet (may append duplicates if odd)
                df_row.to_excel(writer, sheet_name=sheet_date_str, index=False)
    else:
        # create new file with this sheet
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
            df_row.to_excel(writer, sheet_name=sheet_date_str, index=False)

# -------------------------
# Settings helpers
# -------------------------
def get_setting(key, default=None):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT value FROM settings WHERE key = ?", (key,))
    r = cur.fetchone()
    conn.close()
    if r:
        return r['value']
    return default

def set_setting(key, value):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, str(value)))
    conn.commit()
    conn.close()

# -------------------------
# Routes & Views
# -------------------------
INDEX_HTML = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>School Sport Booking</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body class="bg-light">
<div class="container py-4">
  <div class="d-flex justify-content-between align-items-center">
    <h1>School Sport Booking</h1>
    <div>
      <a href="/admin" class="btn btn-outline-secondary btn-sm">Admin</a>
    </div>
  </div>
  <p class="lead">Scan the QR code to open this form on your phone. Book the slot for your sport (only 1 booking per slot).</p>
  <div class="card p-3 mb-3">
    <h5>Quick booking</h5>
    <form id="quickForm" method="GET" action="/book">
      <div class="mb-2">
        <label for="date" class="form-label">Booking date</label>
        <input class="form-control" type="date" id="date" name="date" value="{{today}}">
      </div>
      <button class="btn btn-primary">Open Booking Form</button>
    </form>
  </div>
  <div class="card p-3">
    <h5>Rules</h5>
    <ul>
      <li>One booking per slot (sport + period + date).</li>
      <li>Check the period timings on the booking page.</li>
      <li>Contact the supervisor if you get a message to visit them.</li>
    </ul>
  </div>
</div>
</body>
</html>
"""

BOOK_HTML = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Book a Sport Slot</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
  <script>
  async function checkSlot() {
    const form = document.getElementById('bookForm');
    const data = new URLSearchParams(new FormData(form));
    const res = await fetch('/api/check_slot', {method: 'POST', body: data});
    const j = await res.json();
    const el = document.getElementById('slotInfo');
    if (j.ok) {
      el.innerHTML = '<div class="alert alert-success">Slot available</div>';
    } else {
      el.innerHTML = '<div class="alert alert-danger">Slot not available: ' + j.msg + '</div>';
    }
  }
  </script>
</head>
<body class="bg-light">
<div class="container py-4">
  <h2>Book a sport slot for {{booking_date}}</h2>
  <div class="card p-3">
    <form id="bookForm" method="POST" action="/submit_booking">
      <input type="hidden" name="booking_date" value="{{booking_date}}">
      <div class="mb-2">
        <label for="period" class="form-label">Period</label>
        <select name="period" id="period" class="form-select" onchange="checkSlot()">
          {% for p in periods %}
            <option value="{{p.period}}" data-bookable="{{p.bookable}}"
              {% if not p.bookable %} disabled {% endif %}>
              Period {{p.period}} — {{p.start}} - {{p.end}} {% if not p.bookable %}(not bookable){% endif %}
            </option>
          {% endfor %}
        </select>
      </div>
      <div class="mb-2">
        <label for="sport" class="form-label">Sport</label>
        <select name="sport" class="form-select" onchange="checkSlot()">
          <option>Table Tennis</option>
          <option>Badminton</option>
          <option>Basketball</option>
          <option>Football</option>
          <option>Other</option>
        </select>
      </div>
      <div class="mb-2">
        <label for="name" class="form-label">Your full name</label>
        <input name="name" required class="form-control">
      </div>
      <div class="mb-2">
        <label for="email" class="form-label">School email (optional)</label>
        <input name="email" class="form-control" placeholder="aaisp000000@alansarschool.net">
      </div>
      <div class="mb-2">
        <label for="grade" class="form-label">Grade (Year 12 or Year 13 — many formats allowed)</label>
        <input name="grade" class="form-control" placeholder="e.g. 12, Y12, Grade 11">
      </div>
      <div class="mb-2">
        <label for="others" class="form-label">Other players (comma separated)</label>
        <input name="others" class="form-control" placeholder="Name1, Name2">
      </div>

      <div id="slotInfo"></div>

      <button class="btn btn-primary">Submit booking</button>
    </form>
  </div>

  <hr>
  <a href="/" class="btn btn-link">Back</a>
</div>
</body>
</html>
"""

SUCCESS_HTML = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Booked</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body class="bg-light">
<div class="container py-4">
  <div class="alert alert-success">
    <h4>Booking successful</h4>
    <p>{{msg}}</p>
  </div>
  <a href="/" class="btn btn-primary">Home</a>
</div>
</body>
</html>
"""

FAIL_HTML = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Booking error</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body class="bg-light">
<div class="container py-4">
  <div class="alert alert-danger">
    <h4>Booking failed</h4>
    <p>{{msg}}</p>
  </div>
  <a href="/" class="btn btn-primary">Home</a>
</div>
</body>
</html>
"""

ADMIN_LOGIN_HTML = """
<!doctype html>
<html lang="en">
<head><meta charset="utf-8"><title>Admin Login</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body class="bg-light">
<div class="container py-4">
  <h2>Admin Login</h2>
  <form method="POST" action="/admin/login">
    <div class="mb-2"><input name="username" placeholder="Username" class="form-control"></div>
    <div class="mb-2"><input name="password" type="password" placeholder="Password" class="form-control"></div>
    <button class="btn btn-primary">Login</button>
  </form>
  <a href="/" class="btn btn-link mt-2">Home</a>
</div>
</body>
</html>
"""

ADMIN_DASH_HTML = """
<!doctype html>
<html lang="en">
<head><meta charset="utf-8"><title>Admin</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body class="bg-light">
<div class="container py-4">
  <h2>Supervisor Dashboard</h2>

  <div class="card p-3 mb-3">
    <h5>Settings</h5>
    <form method="POST" action="/admin/settings">
      <div class="row">
        <div class="col">
          <label>Black points threshold</label>
          <input name="blackpoint_threshold" value="{{threshold}}" class="form-control">
        </div>
      </div>
      <button class="btn btn-primary btn-sm mt-2">Save</button>
    </form>
  </div>

  <div class="card p-3 mb-3">
    <h5>Today's bookings ({{today}})</h5>
    {% if bookings %}
      <table class="table">
        <thead><tr><th>ID</th><th>Period</th><th>Sport</th><th>Leader</th><th>Others</th><th>Actions</th></tr></thead>
        <tbody>
        {% for b in bookings %}
          <tr>
            <td>{{b['id']}}</td>
            <td>{{b['period']}}</td>
            <td>{{b['sport']}}</td>
            <td>{{b['leader_name']}}</td>
            <td>{{b['other_players']}}</td>
            <td>
              <form style="display:inline" method="POST" action="/admin/delete/{{b['id']}}">
                <button class="btn btn-danger btn-sm">Delete</button>
              </form>
              <form style="display:inline" method="POST" action="/admin/redflag/{{b['leader_id']}}">
                <button class="btn btn-warning btn-sm">Red-flag / +1 black point</button>
              </form>
            </td>
          </tr>
        {% endfor %}
        </tbody>
      </table>
    {% else %}
      <p>No bookings today.</p>
    {% endif %}
  </div>

  <a href="/admin/export" class="btn btn-outline-success">Download Excel</a>
  <a href="/" class="btn btn-link">Home</a>
</div>
</body>
</html>
"""

# -------------------------
# Web endpoints
# -------------------------
@app.route("/")
def index():
    today_str = date.today().isoformat()
    return render_template_string(INDEX_HTML, today=today_str)

@app.route("/book")
def book():
    # date query param, default to today
    ds = request.args.get("date")
    try:
        if ds:
            booking_date = datetime.strptime(ds, "%Y-%m-%d").date()
        else:
            booking_date = date.today()
    except Exception:
        booking_date = date.today()
    periods = get_day_periods(booking_date)
    # filter out non-school-days
    if not periods:
        return render_template_string(FAIL_HTML, msg="No school on this date (only Mon-Thu supported).")
    return render_template_string(BOOK_HTML, booking_date=booking_date.isoformat(), periods=periods)

@app.route("/api/check_slot", methods=["POST"])
def api_check_slot():
    booking_date = request.form.get("booking_date")
    period = request.form.get("period")
    sport = request.form.get("sport")
    if not (booking_date and period and sport):
        return jsonify({"ok": False, "msg": "missing fields"})
    # check bookable in timetable
    try:
        bdate = datetime.strptime(booking_date, "%Y-%m-%d").date()
    except:
        return jsonify({"ok": False, "msg": "invalid date"})
    periods = get_day_periods(bdate)
    # find period
    p = next((x for x in periods if x['period'] == int(period)), None)
    if not p:
        return jsonify({"ok": False, "msg": "period not valid on this date"})
    if not p['bookable']:
        return jsonify({"ok": False, "msg": "period not bookable by rules"})
    taken = is_slot_taken(booking_date, int(period), sport)
    if taken:
        return jsonify({"ok": False, "msg": "slot already taken"})
    return jsonify({"ok": True})

@app.route("/submit_booking", methods=["POST"])
def submit_booking():
    booking_date = request.form.get("booking_date")
    period = request.form.get("period")
    sport = request.form.get("sport")
    name = request.form.get("name")
    email = request.form.get("email", "").strip().lower()
    grade_raw = request.form.get("grade", "")
    others = request.form.get("others", "")

    # Basic validation
    if not all([booking_date, period, sport, name]):
        return render_template_string(FAIL_HTML, msg="Missing required fields.")

    try:
        bdate = datetime.strptime(booking_date, "%Y-%m-%d").date()
    except:
        return render_template_string(FAIL_HTML, msg="Invalid date format.")

    # validate period bookable
    periods = get_day_periods(bdate)
    p = next((x for x in periods if x['period'] == int(period)), None)
    if not p or not p['bookable']:
        return render_template_string(FAIL_HTML, msg="Selected period cannot be booked.")

    # email enforcement (optional)
    if ENFORCE_SCHOOL_EMAIL and email:
        if not re.match(SCHOOL_EMAIL_PREFIX, email.lower()):
            return render_template_string(FAIL_HTML, msg="Email not allowed; must be school email.")

    # normalize grade
    grade = normalize_grade_input(grade_raw)

    # create/find student
    leader_id = find_or_create_student(name, email, grade)

    # create booking; DB unique constraint ensures only one booking per slot
    success, err = create_booking(booking_date, int(period), sport, leader_id, others)
    if success:
        # check blackpoints threshold
        threshold = int(get_setting("blackpoint_threshold", BLACKPOINTS_DEFAULT_THRESHOLD))
        # fetch leader black points
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT black_points FROM students WHERE id = ?", (leader_id,))
        row = cur.fetchone()
        conn.close()
        leader_bp = row['black_points'] if row else 0
        if leader_bp >= threshold:
            # message: go to supervisor
            return render_template_string(SUCCESS_HTML, msg="Booked successfully. You have been flagged: please see the supervisor in person.")
        return render_template_string(SUCCESS_HTML, msg="Booked successfully. See you at the court!")
    else:
        if err == "slot_taken":
            return render_template_string(FAIL_HTML, msg="Slot already taken. Try a different period or sport.")
        return render_template_string(FAIL_HTML, msg=f"Failed: {err}")

# -------------------------
# Admin routes - simple auth
# -------------------------
from functools import wraps
from flask import session

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("admin_logged_in"):
            return f(*args, **kwargs)
        return redirect(url_for("admin_login"))
    return decorated

@app.route("/admin")
def admin_root():
    return redirect(url_for("admin_login"))

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        u = request.form.get("username")
        p = request.form.get("password")
        if u == ADMIN_USER and p == ADMIN_PASS:
            session["admin_logged_in"] = True
            return redirect(url_for("admin_dashboard"))
        else:
            return render_template_string(ADMIN_LOGIN_HTML)
    return render_template_string(ADMIN_LOGIN_HTML)

@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_logged_in", None)
    return redirect(url_for("index"))

@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():
    today_str = date.today().isoformat()
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT b.id, b.period, b.sport, b.other_players, b.created_at, s.name as leader_name, s.id as leader_id
        FROM bookings b LEFT JOIN students s ON b.leader_student_id = s.id
        WHERE b.booking_date = ?
        ORDER BY b.period
    """, (today_str,))
    rows = cur.fetchall()
    bookings = [dict(r) for r in rows]
    cur.execute("SELECT value FROM settings WHERE key = 'blackpoint_threshold'")
    thr = cur.fetchone()
    threshold = int(thr['value']) if thr else BLACKPOINTS_DEFAULT_THRESHOLD
    conn.close()
    return render_template_string(ADMIN_DASH_HTML, bookings=bookings, today=today_str, threshold=threshold)

@app.route("/admin/delete/<int:bid>", methods=["POST"])
@admin_required
def admin_delete(bid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM bookings WHERE id = ?", (bid,))
    conn.commit()
    conn.close()
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/redflag/<int:student_id>", methods=["POST"])
@admin_required
def admin_redflag(student_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE students SET black_points = black_points + 1 WHERE id = ?", (student_id,))
    conn.commit()
    conn.close()
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/settings", methods=["POST"])
@admin_required
def admin_settings():
    thr = request.form.get("blackpoint_threshold")
    if thr and thr.isdigit():
        set_setting("blackpoint_threshold", int(thr))
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/export")
@admin_required
def admin_export():
    # return the bookings Excel file
    if os.path.exists(EXCEL_FILE):
        return send_file(EXCEL_FILE, as_attachment=True)
    else:
        return "No Excel file yet. First booking will create it."

# -------------------------
# Startup
# -------------------------
if __name__ == "__main__":
    app.run(debug=True)
